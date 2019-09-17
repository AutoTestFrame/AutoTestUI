# -*- coding: utf-8 -*-
import openpyxl
import xlwt,xlrd


class Excel():
    """读写excel类"""
    def __init__(self,filename):
        self.filename = filename

    def read_excel(self, index):
        """
        读取excel文件，返回list格式
        :param index: 第几个sheet页
        :return:
        """
        excle = xlrd.open_workbook(self.filename)
        sheet = excle.sheet_by_index(index) # 第几个sheet页
        rows = sheet.nrows
        cols = sheet.ncols
        headers = sheet.row_values(0) # 第一行数据
        data_list = []
        for i in range(1, rows):    # 第一行是头部信息，所以从第二行开始循环取值
            case_info = {}
            for j in range(0, cols):    # 获取第一列的值
                case_info[headers[j]] = sheet.row_values(i)[j]  # headers的零个值与第一行第零列组成一个字典
            data_list.append(case_info)
        return data_list

    def write_excel(self, index, row, col, value):
        """
        写入excel方法
        :param index: 第几个sheet页
        :param row: 第几行
        :param col: 第几列
        :param value: 写入的值
        :return:
        """
        excel = openpyxl.load_workbook(self.filename)
        sheetnames = excel.style_names      # 返回一个list
        sheet = excel[sheetnames][index]    # 写第几个表格
        sheet.cell(row, col, value)
        sheet.save(self.filename)


if __name__=="__main__":
    path = r"F:\PycharmProjects\classnote\data\interface.xls"
    A = Excel(path).read_excel()
    print(A)

